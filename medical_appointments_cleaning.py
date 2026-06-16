import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── 1. Load Data ───────────────────────────────────────────────────────────
df = pd.read_csv('KaggleV2-May-2016.csv')
print("Original Shape:", df.shape)

changes = []

# ─── 2. Rename Columns (fix typos + standardize) ────────────────────────────
rename_map = {
    'PatientId':      'Patient ID',
    'AppointmentID':  'Appointment ID',
    'Gender':         'Gender',
    'ScheduledDay':   'Scheduled Date',
    'AppointmentDay': 'Appointment Date',
    'Age':            'Age',
    'Neighbourhood':  'Neighbourhood',
    'Scholarship':    'Scholarship',
    'Hipertension':   'Hypertension',
    'Diabetes':       'Diabetes',
    'Alcoholism':     'Alcoholism',
    'Handcap':        'Handicap',
    'SMS_received':   'SMS Received',
    'No-show':        'No Show',
}
df.rename(columns=rename_map, inplace=True)
changes.append("Renamed columns: fixed typos ('Hipertension'→'Hypertension', 'Handcap'→'Handicap'), standardized to Title Case")

# ─── 3. Parse & Clean Date Columns ──────────────────────────────────────────
df['Scheduled Date']   = pd.to_datetime(df['Scheduled Date'],   utc=True).dt.date
df['Appointment Date'] = pd.to_datetime(df['Appointment Date'], utc=True).dt.date
changes.append("Standardized dates: stripped ISO 8601 timestamps to date-only (YYYY-MM-DD)")

# ─── 4. Remove Negative Age ─────────────────────────────────────────────────
neg_age = (df['Age'] < 0).sum()
df = df[df['Age'] >= 0].copy()
changes.append(f"Removed {neg_age} row(s) with invalid negative Age value")

# ─── 5. Remove Rows Where Scheduled Date > Appointment Date ─────────────────
bad_dates = (df['Scheduled Date'] > df['Appointment Date']).sum()
df = df[df['Scheduled Date'] <= df['Appointment Date']].copy()
changes.append(f"Removed {bad_dates} rows where Scheduled Date was after Appointment Date (logically invalid)")

# ─── 6. Fix Handicap Column (binary flag, clamp values > 1) ─────────────────
hcap_fixed = (df['Handicap'] > 1).sum()
df['Handicap'] = df['Handicap'].clip(upper=1)
changes.append(f"Clamped {hcap_fixed} Handicap values > 1 to 1 (binary flag column)")

# ─── 7. Expand Gender Codes ─────────────────────────────────────────────────
df['Gender'] = df['Gender'].map({'F': 'Female', 'M': 'Male'})
changes.append("Expanded Gender codes: 'F'→'Female', 'M'→'Male'")

# ─── 8. Encode No Show as 0/1 ───────────────────────────────────────────────
df['No Show'] = df['No Show'].map({'No': 0, 'Yes': 1})
changes.append("Encoded 'No Show': 'No'→0, 'Yes'→1")

# ─── 9. Standardize Neighbourhood ───────────────────────────────────────────
df['Neighbourhood'] = df['Neighbourhood'].str.strip().str.upper()
changes.append("Standardized Neighbourhood to UPPERCASE and stripped whitespace")

# ─── 10. Fix Patient ID (float → integer string) ────────────────────────────
df['Patient ID'] = df['Patient ID'].apply(lambda x: str(int(x)) if pd.notna(x) else '')
changes.append("Converted Patient ID from float64 to integer string")

# ─── 11. Duplicate Check ────────────────────────────────────────────────────
dups = df.duplicated().sum()
changes.append(f"Checked duplicates: {dups} full duplicate rows found (none removed)")

print("Cleaned Shape:", df.shape)
print("\nChanges Made:")
for i, c in enumerate(changes, 1):
    print(f"  {i}. {c}")

# ─── 12. Export to Excel ─────────────────────────────────────────────────────
output_path = 'medical_appointments_cleaned.xlsx'
df.to_excel(output_path, index=False, sheet_name='Cleaned Data')

# ─── 13. Apply Excel Formatting ─────────────────────────────────────────────
wb = load_workbook(output_path)
ws = wb['Cleaned Data']

HDR_FILL  = PatternFill('solid', start_color='2E86AB')
ALT_FILL  = PatternFill('solid', start_color='EEF4F8')
HDR_FONT  = Font(name='Arial', bold=True, color='FFFFFF', size=10)
BODY_FONT = Font(name='Arial', size=10)
CENTER    = Alignment(horizontal='center', vertical='center')
LEFT      = Alignment(horizontal='left',   vertical='center')
THIN      = Border(
    bottom=Side(style='thin', color='C9D8E3'),
    right =Side(style='thin', color='C9D8E3')
)

col_widths = [18, 16, 10, 15, 17, 6, 22, 11, 12, 10, 12, 10, 13, 9]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

for cell in ws[1]:
    cell.font      = HDR_FONT
    cell.fill      = HDR_FILL
    cell.alignment = CENTER
    cell.border    = THIN

for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
    fill = ALT_FILL if row_idx % 2 == 0 else PatternFill()
    for cell in row:
        cell.font      = BODY_FONT
        cell.fill      = fill
        cell.border    = THIN
        cell.alignment = CENTER

ws.freeze_panes = 'A2'

# ─── 14. Add Cleaning Summary Sheet ─────────────────────────────────────────
ws2 = wb.create_sheet('Cleaning Summary')
ws2.column_dimensions['A'].width = 5
ws2.column_dimensions['B'].width = 90

ws2['A1'] = '#'
ws2['B1'] = 'Cleaning Action Performed'
for cell in ws2['1:1']:
    cell.font      = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    cell.fill      = HDR_FILL
    cell.alignment = CENTER

for i, ch in enumerate(changes, 1):
    ws2[f'A{i+1}'] = i
    ws2[f'B{i+1}'] = ch
    ws2[f'A{i+1}'].font      = BODY_FONT
    ws2[f'B{i+1}'].font      = BODY_FONT
    ws2[f'A{i+1}'].alignment = CENTER
    ws2[f'B{i+1}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    if i % 2 == 0:
        ws2[f'A{i+1}'].fill = ALT_FILL
        ws2[f'B{i+1}'].fill = ALT_FILL
    ws2.row_dimensions[i+1].height = 32

wb.save(output_path)
print(f"\n✅ File saved: {output_path}")
